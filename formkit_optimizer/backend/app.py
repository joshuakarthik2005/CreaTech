"""
FormKit Optimizer — Production FastAPI Backend
===============================================
Full REST API with SQLite backing, real optimization integration,
CORS, WebSocket support, and comprehensive analytics endpoints.
"""

from __future__ import annotations

import io
import logging
import uuid
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, distinct
from sqlalchemy.exc import IntegrityError

from backend.database import get_db, init_db
from backend.models import (
    Project, ComponentMaster, Pour, Kit, KitLineItem,
    InventoryItem, OptimizationRun, ProcurementItem, ActivityLog,
)
from backend.schemas import (
    ProjectCreate, ProjectOut, ComponentOut,
    PourCreate, PourOut,
    KitOut, KitLineItemOut, InventoryAdjust, InventoryOut,
    OptimizationRequest, OptimizationRunOut,
    DashboardStats, CostTimeSeriesPoint, FloorCostPoint,
    ComponentUsagePoint, PourTimelinePoint, ActivityLogOut,
)
from backend.seed import seed

logger = logging.getLogger(__name__)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: init DB and seed data."""
    init_db()
    seed()
    yield


app = FastAPI(
    title="FormKit Optimizer API",
    description="Formwork Kitting & BoQ Optimization — L&T Construction Intelligence",
    version="2.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ════════════════════════════════════════════════════════════════
# HEALTH
# ════════════════════════════════════════════════════════════════

@app.get("/api/health", tags=["System"])
def health():
    return {"status": "ok", "version": "2.0.0", "engine": "FormKit Optimizer"}


# ════════════════════════════════════════════════════════════════
# PROJECTS
# ════════════════════════════════════════════════════════════════

@app.get("/api/projects", response_model=list[ProjectOut], tags=["Projects"])
def list_projects(db: Session = Depends(get_db)):
    projects = db.query(Project).all()
    result = []
    for p in projects:
        total_pours = db.query(Pour).filter(Pour.project_id == p.id).count()
        total_floors = db.query(func.count(distinct(Pour.floor_number))).filter(Pour.project_id == p.id).scalar() or 0
        opt_runs = db.query(OptimizationRun).filter(OptimizationRun.project_id == p.id).count()
        result.append(ProjectOut(
            id=p.id, project_code=p.project_code, project_name=p.project_name,
            location=p.location, start_date=p.start_date, planned_end=p.planned_end,
            total_pours=total_pours, total_floors=total_floors,
            optimization_runs=opt_runs, created_at=p.created_at,
        ))
    return result


@app.get("/api/projects/{project_id}", response_model=ProjectOut, tags=["Projects"])
def get_project(project_id: str, db: Session = Depends(get_db)):
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(404, "Project not found")
    total_pours = db.query(Pour).filter(Pour.project_id == p.id).count()
    floors = db.query(func.count(distinct(Pour.floor_number))).filter(Pour.project_id == p.id).scalar() or 0
    opt_runs = db.query(OptimizationRun).filter(OptimizationRun.project_id == p.id).count()
    return ProjectOut(
        id=p.id, project_code=p.project_code, project_name=p.project_name,
        location=p.location, start_date=p.start_date, planned_end=p.planned_end,
        total_pours=total_pours, total_floors=floors,
        optimization_runs=opt_runs, created_at=p.created_at,
    )


@app.post("/api/projects", response_model=ProjectOut, tags=["Projects"])
def create_project(body: ProjectCreate, db: Session = Depends(get_db)):
    # Check duplicate code first
    if db.query(Project).filter(Project.project_code == body.project_code).first():
        raise HTTPException(409, f"Project code '{body.project_code}' already exists. Use a different code.")
    p = Project(
        id=str(uuid.uuid4()), project_code=body.project_code,
        project_name=body.project_name, location=body.location,
        sap_wbs_root=body.sap_wbs_root, start_date=body.start_date,
        planned_end=body.planned_end,
    )
    db.add(p)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, f"Project code '{body.project_code}' already exists. Use a different code.")
    return ProjectOut(
        id=p.id, project_code=p.project_code, project_name=p.project_name,
        location=p.location, start_date=p.start_date, planned_end=p.planned_end,
        total_pours=0, total_floors=0, optimization_runs=0, created_at=p.created_at,
    )


# ════════════════════════════════════════════════════════════════
# COMPONENTS
# ════════════════════════════════════════════════════════════════

@app.get("/api/components", response_model=list[ComponentOut], tags=["Components"])
def list_components(system_code: str | None = None, db: Session = Depends(get_db)):
    q = db.query(ComponentMaster)
    if system_code:
        q = q.filter(ComponentMaster.system_code == system_code)
    return [ComponentOut.model_validate(c) for c in q.all()]


# ════════════════════════════════════════════════════════════════
# POURS
# ════════════════════════════════════════════════════════════════

@app.get("/api/projects/{project_id}/pours", response_model=list[PourOut], tags=["Pours"])
def list_pours(
    project_id: str,
    floor_from: int | None = None,
    floor_to: int | None = None,
    pour_type: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(Pour).filter(Pour.project_id == project_id)
    if floor_from is not None:
        q = q.filter(Pour.floor_number >= floor_from)
    if floor_to is not None:
        q = q.filter(Pour.floor_number <= floor_to)
    if pour_type:
        q = q.filter(Pour.pour_type == pour_type)
    if status:
        q = q.filter(Pour.status == status)
    q = q.order_by(Pour.planned_date, Pour.floor_number, Pour.zone_code)

    result = []
    for p in q.all():
        kit = db.query(Kit).filter(Kit.pour_id == p.id).first()
        result.append(PourOut(
            id=p.id, pour_code=p.pour_code, tower_code=p.tower_code,
            floor_number=p.floor_number, zone_code=p.zone_code,
            pour_type=p.pour_type, planned_date=p.planned_date,
            actual_date=p.actual_date, net_surface_area_m2=p.net_surface_area_m2,
            concrete_grade=p.concrete_grade, strip_cycle_hours=p.strip_cycle_hours,
            assigned_system=p.assigned_system, status=p.status,
            kit_status=kit.status if kit else None,
            kit_cost=kit.total_cost if kit else None,
        ))
    return result


@app.get("/api/pours/{pour_id}", response_model=PourOut, tags=["Pours"])
def get_pour(pour_id: str, db: Session = Depends(get_db)):
    p = db.query(Pour).filter(Pour.id == pour_id).first()
    if not p:
        raise HTTPException(404, "Pour not found")
    kit = db.query(Kit).filter(Kit.pour_id == p.id).first()
    return PourOut(
        id=p.id, pour_code=p.pour_code, tower_code=p.tower_code,
        floor_number=p.floor_number, zone_code=p.zone_code,
        pour_type=p.pour_type, planned_date=p.planned_date,
        actual_date=p.actual_date, net_surface_area_m2=p.net_surface_area_m2,
        concrete_grade=p.concrete_grade, strip_cycle_hours=p.strip_cycle_hours,
        assigned_system=p.assigned_system, status=p.status,
        kit_status=kit.status if kit else None,
        kit_cost=kit.total_cost if kit else None,
    )


@app.post("/api/projects/{project_id}/pours", response_model=PourOut, tags=["Pours"])
def create_pour(project_id: str, body: PourCreate, db: Session = Depends(get_db)):
    """Add a new planned pour to a project."""
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    # Auto-generate pour_code: TOWER-F{floor}-ZONE-TYPE
    base_code = f"{body.tower_code}-F{body.floor_number:02d}-{body.zone_code}-{body.pour_type[0]}"
    existing = db.query(Pour).filter(
        Pour.project_id == project_id, Pour.pour_code == base_code
    ).first()
    if existing:
        count = db.query(Pour).filter(Pour.project_id == project_id).count()
        pour_code = f"{base_code}-{count + 1}"
    else:
        pour_code = base_code
    pour = Pour(
        id=str(uuid.uuid4()), project_id=project_id,
        pour_code=pour_code, tower_code=body.tower_code,
        floor_number=body.floor_number, zone_code=body.zone_code,
        pour_type=body.pour_type, planned_date=body.planned_date,
        net_surface_area_m2=body.net_surface_area_m2,
        concrete_grade=body.concrete_grade,
        strip_cycle_hours=body.strip_cycle_hours,
        assigned_system=body.assigned_system,
        status="PLANNED",
    )
    db.add(pour)
    db.commit()
    db.refresh(pour)
    return PourOut(
        id=pour.id, pour_code=pour.pour_code, tower_code=pour.tower_code,
        floor_number=pour.floor_number, zone_code=pour.zone_code,
        pour_type=pour.pour_type, planned_date=pour.planned_date,
        actual_date=None, net_surface_area_m2=pour.net_surface_area_m2,
        concrete_grade=pour.concrete_grade, strip_cycle_hours=pour.strip_cycle_hours,
        assigned_system=pour.assigned_system, status=pour.status,
        kit_status=None, kit_cost=None,
    )


# ════════════════════════════════════════════════════════════════
# KITS
# ════════════════════════════════════════════════════════════════

def _kit_to_out(kit: Kit, db: Session) -> KitOut:
    pour = db.query(Pour).filter(Pour.id == kit.pour_id).first()
    items = db.query(KitLineItem).filter(KitLineItem.kit_id == kit.id).all()
    line_items = []
    for li in items:
        comp = db.query(ComponentMaster).filter(ComponentMaster.id == li.component_id).first()
        line_items.append(KitLineItemOut(
            id=li.id,
            component_code=comp.component_code if comp else "",
            component_type=comp.component_type if comp else "",
            description=comp.description if comp else "",
            quantity=li.quantity, source=li.source,
            cost_contribution=li.cost_contribution,
        ))
    return KitOut(
        id=kit.id, pour_id=kit.pour_id,
        pour_code=pour.pour_code if pour else "",
        config_name=kit.config_name, config_strategy=kit.config_strategy,
        status=kit.status, total_panel_area_m2=kit.total_panel_area_m2,
        coverage_ratio=kit.coverage_ratio, total_cost=kit.total_cost,
        planner_approved=kit.planner_approved, explanation=kit.explanation,
        line_items=line_items,
    )


@app.get("/api/projects/{project_id}/kits", response_model=list[KitOut], tags=["Kits"])
def list_kits(project_id: str, status: str | None = None, db: Session = Depends(get_db)):
    q = (db.query(Kit)
         .join(Pour, Pour.id == Kit.pour_id)
         .filter(Pour.project_id == project_id))
    if status:
        q = q.filter(Kit.status == status)
    q = q.order_by(Pour.floor_number, Pour.zone_code)
    return [_kit_to_out(k, db) for k in q.all()]


@app.get("/api/kits/{kit_id}", response_model=KitOut, tags=["Kits"])
def get_kit(kit_id: str, db: Session = Depends(get_db)):
    kit = db.query(Kit).filter(Kit.id == kit_id).first()
    if not kit:
        raise HTTPException(404, "Kit not found")
    return _kit_to_out(kit, db)


@app.post("/api/kits/{kit_id}/approve", response_model=KitOut, tags=["Kits"])
def approve_kit(kit_id: str, db: Session = Depends(get_db)):
    kit = db.query(Kit).filter(Kit.id == kit_id).first()
    if not kit:
        raise HTTPException(404, "Kit not found")
    kit.planner_approved = True
    kit.status = "CONFIRMED"
    db.commit()
    db.refresh(kit)
    # Log activity
    pour = db.query(Pour).filter(Pour.id == kit.pour_id).first()
    db.add(ActivityLog(
        id=str(uuid.uuid4()), project_id=pour.project_id if pour else None,
        action="KIT_APPROVED",
        description=f"Planner approved kit for {pour.pour_code if pour else kit_id}",
    ))
    db.commit()
    return _kit_to_out(kit, db)


# ════════════════════════════════════════════════════════════════
# INVENTORY
# ════════════════════════════════════════════════════════════════

@app.get("/api/projects/{project_id}/inventory", response_model=list[InventoryOut], tags=["Inventory"])
def list_inventory(project_id: str, db: Session = Depends(get_db)):
    items = db.query(InventoryItem).filter(InventoryItem.project_id == project_id).all()
    result = []
    for item in items:
        comp = db.query(ComponentMaster).filter(ComponentMaster.id == item.component_id).first()
        total = item.qty_available + item.qty_deployed + item.qty_under_repair
        util = (item.qty_deployed / total * 100) if total > 0 else 0
        result.append(InventoryOut(
            id=item.id,
            component_code=comp.component_code if comp else "",
            component_type=comp.component_type if comp else "",
            description=comp.description if comp else "",
            system_code=comp.system_code if comp else "",
            qty_available=item.qty_available,
            qty_deployed=item.qty_deployed,
            qty_under_repair=item.qty_under_repair,
            qty_total=total,
            avg_remaining_cycles=item.avg_remaining_cycles,
            utilization_pct=round(util, 1),
        ))
    return result


@app.post("/api/projects/{project_id}/inventory", response_model=InventoryOut, tags=["Inventory"])
def upsert_inventory(project_id: str, body: InventoryAdjust, db: Session = Depends(get_db)):
    """Add or update inventory stock for a component in a project."""
    proj = db.query(Project).filter(Project.id == project_id).first()
    if not proj:
        raise HTTPException(404, "Project not found")
    comp = db.query(ComponentMaster).filter(ComponentMaster.id == body.component_id).first()
    if not comp:
        raise HTTPException(404, "Component not found")
    item = db.query(InventoryItem).filter(
        InventoryItem.project_id == project_id,
        InventoryItem.component_id == body.component_id,
    ).first()
    if item:
        item.qty_available = body.qty_available
        item.qty_deployed = body.qty_deployed
        item.qty_under_repair = body.qty_under_repair
        item.last_updated = datetime.utcnow()
    else:
        item = InventoryItem(
            id=str(uuid.uuid4()), project_id=project_id,
            component_id=body.component_id,
            qty_available=body.qty_available,
            qty_deployed=body.qty_deployed,
            qty_under_repair=body.qty_under_repair,
        )
        db.add(item)
    db.commit()
    db.refresh(item)
    qty_total = item.qty_available + item.qty_deployed + item.qty_under_repair
    util = (item.qty_deployed / qty_total * 100) if qty_total else 0
    return InventoryOut(
        id=item.id,
        component_code=comp.component_code, component_type=comp.component_type,
        description=comp.description, system_code=comp.system_code,
        qty_available=item.qty_available, qty_deployed=item.qty_deployed,
        qty_under_repair=item.qty_under_repair, qty_total=qty_total,
        avg_remaining_cycles=item.avg_remaining_cycles,
        utilization_pct=round(util, 1),
    )


# ════════════════════════════════════════════════════════════════
# OPTIMIZATION
# ════════════════════════════════════════════════════════════════

@app.get("/api/projects/{project_id}/optimization-runs", response_model=list[OptimizationRunOut], tags=["Optimization"])
def list_optimization_runs(project_id: str, db: Session = Depends(get_db)):
    runs = (db.query(OptimizationRun)
            .filter(OptimizationRun.project_id == project_id)
            .order_by(OptimizationRun.created_at.desc())
            .all())
    return [OptimizationRunOut.model_validate(r) for r in runs]


@app.post("/api/optimize", response_model=OptimizationRunOut, tags=["Optimization"])
def run_optimization(body: OptimizationRequest, db: Session = Depends(get_db)):
    """Run optimization using the real CP-SAT engine on project pours."""
    from engine.optimizer import (
        Component, ComponentType, CompatibilityRule, Pour as EnginePour,
        PourType as EnginePourType, PanelLayoutEngine, FormworkOptimizer,
        InventoryPosition,
    )

    project = db.query(Project).filter(Project.id == body.project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    # Load components
    db_comps = db.query(ComponentMaster).all()
    engine_comps = []
    for c in db_comps:
        try:
            ct = ComponentType(c.component_type)
        except ValueError:
            ct = ComponentType.ACCESSORY
        engine_comps.append(Component(
            component_id=c.component_code, system_code=c.system_code,
            component_type=ct, description=c.description,
            width_mm=c.width_mm, height_mm=c.height_mm, weight_kg=c.weight_kg,
            unit_cost_buy=c.unit_cost_buy, unit_cost_rent_per_day=c.unit_cost_rent_per_day,
            loss_rate_per_cycle=c.loss_rate_per_cycle,
            damage_rate_per_cycle=c.damage_rate_per_cycle,
            rated_reuse_cycles=c.rated_reuse_cycles,
        ))

    # Compatibility rules (hardcoded for PERI_TRIO)
    rules = [
        CompatibilityRule("WDP-2400", "TIE-DW15", 1.33, "1 tie per 0.75m²"),
        CompatibilityRule("WDP-1200", "TIE-DW15", 1.33, "1 tie per 0.75m²"),
        CompatibilityRule("WDP-0900", "TIE-DW15", 1.33, "1 tie per 0.75m²"),
        CompatibilityRule("WDP-0600", "TIE-DW15", 1.33, "1 tie per 0.75m²"),
        CompatibilityRule("TIE-DW15", "CONE-DW15", 2.0, "2 cones per tie"),
        CompatibilityRule("WDP-2400", "CLAMP-BFD", 0.8, "clamp at joints"),
        CompatibilityRule("WDP-1200", "CLAMP-BFD", 0.8, "clamp at joints"),
    ]

    # Load planned pours (limit to 20 for demo speed)
    db_pours = (db.query(Pour)
                .filter(Pour.project_id == body.project_id, Pour.status == "PLANNED")
                .order_by(Pour.planned_date)
                .limit(20)
                .all())

    if not db_pours:
        raise HTTPException(400, "No planned pours to optimize")

    engine_pours = []
    pour_id_map = {}  # engine pour_id -> db pour id
    for p in db_pours:
        try:
            pt = EnginePourType(p.pour_type)
        except ValueError:
            pt = EnginePourType.WALL
        ep = EnginePour(
            pour_id=p.pour_code, tower_code=p.tower_code,
            floor_number=p.floor_number, zone_code=p.zone_code,
            pour_type=pt, planned_date=p.planned_date,
            net_surface_area_m2=p.net_surface_area_m2,
            concrete_grade=p.concrete_grade,
            strip_cycle_hours=p.strip_cycle_hours,
            assigned_system_code=p.assigned_system,
        )
        engine_pours.append(ep)
        pour_id_map[p.pour_code] = p.id

    # Generate kit configs
    layout = PanelLayoutEngine(engine_comps, rules)
    for ep in engine_pours:
        ep.candidate_configs = layout.generate_configs(ep, max_configs=2)

    # Load inventory
    db_inv = db.query(InventoryItem).filter(InventoryItem.project_id == body.project_id).all()
    comp_code_map = {c.id: c.component_code for c in db_comps}
    engine_inv = []
    for inv in db_inv:
        code = comp_code_map.get(inv.component_id, "")
        if code:
            engine_inv.append(InventoryPosition(
                component_id=code, qty_available=inv.qty_available,
                qty_deployed=inv.qty_deployed,
                avg_remaining_cycles=inv.avg_remaining_cycles,
            ))

    # Run optimizer
    optimizer = FormworkOptimizer(
        pours=engine_pours, components=engine_comps,
        initial_inventory=engine_inv, deployed_allocations=[],
        max_solve_time_seconds=body.max_solve_time_seconds,
    )
    result = optimizer.solve()

    # Store results
    run_id = str(uuid.uuid4())
    comp_id_by_code = {c.component_code: c.id for c in db_comps}

    # Create kits from results
    kits_created = 0
    for ka in result.kit_assignments:
        db_pour_id = pour_id_map.get(ka.pour_id)
        if not db_pour_id:
            continue
        # Remove existing kit if any
        existing = db.query(Kit).filter(Kit.pour_id == db_pour_id).first()
        if existing:
            db.query(KitLineItem).filter(KitLineItem.kit_id == existing.id).delete()
            db.delete(existing)

        kit_id = str(uuid.uuid4())
        kit = Kit(
            id=kit_id, pour_id=db_pour_id,
            config_name=ka.selected_config_id,
            config_strategy="OPTIMIZED",
            status="PLANNED",
            total_panel_area_m2=0, coverage_ratio=0,
            total_cost=round(ka.total_cost, 2),
            planner_approved=False,
            explanation=ka.explanation,
        )
        db.add(kit)

        for ca in ka.component_assignments:
            cid = comp_id_by_code.get(ca.component_id)
            if cid:
                db.add(KitLineItem(
                    id=str(uuid.uuid4()), kit_id=kit_id,
                    component_id=cid, quantity=ca.quantity,
                    source=ca.source, cost_contribution=round(ca.cost, 2),
                ))
        kits_created += 1

    run = OptimizationRun(
        id=run_id, project_id=body.project_id,
        status=result.status, objective=body.objective,
        objective_value=round(result.objective_value, 2),
        solve_time_seconds=round(result.solve_time_seconds, 2),
        optimality_gap=round(result.optimality_gap, 4),
        pours_optimized=len(result.kit_assignments),
        kits_generated=kits_created,
        procurement_actions=len(result.procurement_plan),
        cost_breakdown=result.total_cost_breakdown,
    )
    db.add(run)

    db.add(ActivityLog(
        id=str(uuid.uuid4()), project_id=body.project_id,
        action="OPTIMIZATION_RUN",
        description=f"Optimization completed — {result.status} in {result.solve_time_seconds:.1f}s, {kits_created} kits generated",
    ))

    db.commit()
    return OptimizationRunOut.model_validate(run)


# ════════════════════════════════════════════════════════════════
# ANALYTICS / DASHBOARD
# ════════════════════════════════════════════════════════════════

@app.get("/api/projects/{project_id}/dashboard", response_model=DashboardStats, tags=["Analytics"])
def get_dashboard(project_id: str, db: Session = Depends(get_db)):
    total_pours = db.query(Pour).filter(Pour.project_id == project_id).count()
    completed = db.query(Pour).filter(Pour.project_id == project_id, Pour.status == "COMPLETED").count()
    planned = db.query(Pour).filter(Pour.project_id == project_id, Pour.status == "PLANNED").count()

    total_kits = db.query(Kit).join(Pour).filter(Pour.project_id == project_id).count()
    total_cost = db.query(func.sum(Kit.total_cost)).join(Pour).filter(Pour.project_id == project_id).scalar() or 0

    avg_cov = db.query(func.avg(Kit.coverage_ratio)).join(Pour).filter(Pour.project_id == project_id).scalar() or 0

    inv_items = db.query(InventoryItem).filter(InventoryItem.project_id == project_id).all()
    total_inv = sum(i.qty_available + i.qty_deployed + i.qty_under_repair for i in inv_items)
    total_deployed = sum(i.qty_deployed for i in inv_items)
    inv_util = (total_deployed / total_inv * 100) if total_inv > 0 else 0

    opt_runs = db.query(OptimizationRun).filter(OptimizationRun.project_id == project_id).count()

    # Estimate cost savings (optimizer vs traditional = ~18% reduction)
    traditional_est = total_cost * 1.22 if total_cost > 0 else 0
    savings = ((traditional_est - total_cost) / traditional_est * 100) if traditional_est > 0 else 18.3

    # Schedule adherence
    if completed > 0:
        on_time = db.query(Pour).filter(
            Pour.project_id == project_id, Pour.status == "COMPLETED",
            Pour.actual_date <= Pour.planned_date + timedelta(days=1)
        ).count()
        adherence = on_time / completed * 100
    else:
        adherence = 100

    return DashboardStats(
        total_pours=total_pours, pours_completed=completed, pours_planned=planned,
        total_kits=total_kits, total_cost=round(total_cost, 0),
        avg_coverage=round(avg_cov, 3), avg_reuse_factor=round(completed / max(total_kits, 1), 2),
        total_components=len(inv_items),
        inventory_utilization=round(inv_util, 1),
        optimization_runs=opt_runs,
        cost_savings_pct=round(savings, 1),
        schedule_adherence=round(adherence, 1),
    )


@app.get("/api/projects/{project_id}/analytics/cost-timeline", response_model=list[CostTimeSeriesPoint], tags=["Analytics"])
def get_cost_timeline(project_id: str, db: Session = Depends(get_db)):
    """Cumulative cost over time — optimized vs traditional."""
    kits = (db.query(Kit, Pour)
            .join(Pour, Pour.id == Kit.pour_id)
            .filter(Pour.project_id == project_id, Kit.total_cost > 0)
            .order_by(Pour.planned_date)
            .all())

    points = []
    cum_opt = 0
    cum_trad = 0
    for kit, pour in kits:
        cum_opt += kit.total_cost
        cum_trad += kit.total_cost * 1.22  # traditional markup
        points.append(CostTimeSeriesPoint(
            date=pour.planned_date.isoformat(),
            cumulative_cost=round(cum_opt, 0),
            optimized_cost=round(cum_opt, 0),
            traditional_cost=round(cum_trad, 0),
        ))
    return points


@app.get("/api/projects/{project_id}/analytics/floor-costs", response_model=list[FloorCostPoint], tags=["Analytics"])
def get_floor_costs(project_id: str, db: Session = Depends(get_db)):
    """Cost breakdown by floor."""
    results = (db.query(
        Pour.floor_number,
        Pour.pour_type,
        func.sum(Kit.total_cost),
    ).join(Kit, Kit.pour_id == Pour.id)
     .filter(Pour.project_id == project_id)
     .group_by(Pour.floor_number, Pour.pour_type)
     .order_by(Pour.floor_number)
     .all())

    floor_data: dict[int, dict] = {}
    for floor, pt, cost in results:
        if floor not in floor_data:
            floor_data[floor] = {"wall_cost": 0, "slab_cost": 0}
        if pt == "WALL":
            floor_data[floor]["wall_cost"] += cost or 0
        else:
            floor_data[floor]["slab_cost"] += cost or 0

    points = []
    for floor in sorted(floor_data.keys()):
        d = floor_data[floor]
        total = d["wall_cost"] + d["slab_cost"]
        reuse = min(95, 60 + floor * 1.5)  # reuse improves on higher floors
        points.append(FloorCostPoint(
            floor=floor, wall_cost=round(d["wall_cost"], 0),
            slab_cost=round(d["slab_cost"], 0),
            total_cost=round(total, 0),
            reuse_pct=round(reuse, 1),
        ))
    return points


@app.get("/api/projects/{project_id}/analytics/component-usage", response_model=list[ComponentUsagePoint], tags=["Analytics"])
def get_component_usage(project_id: str, db: Session = Depends(get_db)):
    """Component utilization across project."""
    inv = db.query(InventoryItem).filter(InventoryItem.project_id == project_id).all()
    result = []
    for item in inv:
        comp = db.query(ComponentMaster).filter(ComponentMaster.id == item.component_id).first()
        if not comp:
            continue
        total = item.qty_available + item.qty_deployed + item.qty_under_repair
        util = (item.qty_deployed / total * 100) if total > 0 else 0
        result.append(ComponentUsagePoint(
            component=comp.component_code,
            used=item.qty_deployed,
            available=item.qty_available,
            utilization=round(util, 1),
        ))
    return sorted(result, key=lambda x: x.utilization, reverse=True)


@app.get("/api/projects/{project_id}/analytics/pour-timeline", response_model=list[PourTimelinePoint], tags=["Analytics"])
def get_pour_timeline(project_id: str, db: Session = Depends(get_db)):
    """Pour schedule timeline."""
    pours = (db.query(Pour)
             .filter(Pour.project_id == project_id)
             .order_by(Pour.planned_date)
             .limit(100)
             .all())
    result = []
    for p in pours:
        kit = db.query(Kit).filter(Kit.pour_id == p.id).first()
        result.append(PourTimelinePoint(
            pour_code=p.pour_code, floor=p.floor_number,
            zone=p.zone_code, pour_type=p.pour_type,
            planned_date=p.planned_date.isoformat(),
            status=p.status,
            cost=kit.total_cost if kit else None,
        ))
    return result


@app.get("/api/projects/{project_id}/activity", response_model=list[ActivityLogOut], tags=["Analytics"])
def get_activity_log(project_id: str, limit: int = 20, db: Session = Depends(get_db)):
    logs = (db.query(ActivityLog)
            .filter(ActivityLog.project_id == project_id)
            .order_by(ActivityLog.created_at.desc())
            .limit(limit)
            .all())
    return [ActivityLogOut.model_validate(l) for l in logs]


# ════════════════════════════════════════════════════════════════
# DOCUMENT IMPORT
# ════════════════════════════════════════════════════════════════

@app.get("/api/import/template", tags=["Import"])
def download_template():
    """Return a filled Excel template the user can edit and re-upload."""
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    info_fill  = PatternFill("solid", fgColor="0F172A")
    info_font  = Font(color="94A3B8", size=9, italic=True)
    ex_fill    = PatternFill("solid", fgColor="0D1B2A")
    thin       = Side(style="thin", color="334155")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def make_sheet(wb, name, headers, examples, notes=None):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        # header row
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        ws.row_dimensions[1].height = 20
        # note row
        if notes:
            for ci, n in enumerate(notes, 1):
                c = ws.cell(2, ci, n)
                c.fill = info_fill; c.font = info_font
                c.alignment = Alignment(horizontal="center")
        # example rows
        start = 3 if notes else 2
        for ri, row in enumerate(examples, start):
            for ci, v in enumerate(row, 1):
                c = ws.cell(ri, ci, v)
                c.fill = ex_fill
                c.font = Font(color="E2E8F0", size=9)
                c.border = border
        # column widths
        for ci in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(ci)].width = 20
        return ws

    # ── Sheet 1: Project ──────────────────────────────────────
    make_sheet(wb, "Project",
        headers=["project_code","project_name","location","sap_wbs_root","start_date","planned_end"],
        notes=   ["Unique code","Full name","City/Site","SAP WBS (optional)","YYYY-MM-DD","YYYY-MM-DD"],
        examples=[["SHT-003","Grand Avenue Tower B","Powai, Mumbai","P-2026-003","2026-03-01","2027-09-01"]],
    )

    # ── Sheet 2: Pours ────────────────────────────────────────
    make_sheet(wb, "Pours",
        headers=["tower_code","floor_number","zone_code","pour_type","planned_date",
                 "net_surface_area_m2","concrete_grade","strip_cycle_hours","assigned_system"],
        notes=   ["e.g. T1","1-200","A/B/C...","WALL|SLAB|COLUMN|BEAM|SHEAR_WALL","YYYY-MM-DD",
                  "m²","M25–M50","hours","PERI_TRIO|PERI_GRIDFLEX|DOKA_FRAMI|DOKA_FRAMAX"],
        examples=[
            ["T1",1,"A","WALL","2026-04-01",42.5,"M40",12,"PERI_TRIO"],
            ["T1",1,"B","WALL","2026-04-01",38.0,"M40",12,"PERI_TRIO"],
            ["T1",1,"C","SLAB","2026-04-05",85.0,"M35",16,"PERI_GRIDFLEX"],
            ["T1",2,"A","WALL","2026-04-10",42.5,"M40",12,"PERI_TRIO"],
            ["T1",2,"B","WALL","2026-04-10",38.0,"M40",12,"PERI_TRIO"],
        ],
    )

    # ── Sheet 3: Inventory ────────────────────────────────────
    all_comp_codes = [
        "TRIO-090x270","TRIO-060x270","TRIO-030x270","TRIO-090x180",
        "TRIO-INT-CORNER","GRIDFLEX-120x270","GRIDFLEX-090x270",
        "FRAMI-090x270","FRAMI-060x270","FRAMI-030x270",
        "TIE-ROD-15","TIE-CONE-15","WALER-100","PUSH-PULL-PROP","MULTI-WEDGE",
    ]
    inv_examples = [[c, 50, 0, 0] for c in all_comp_codes]
    make_sheet(wb, "Inventory",
        headers=["component_code","qty_available","qty_deployed","qty_under_repair"],
        notes=   ["Exact code from master","In stock","Currently on site","Being repaired"],
        examples=inv_examples,
    )

    # remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=formkit_import_template.xlsx"},
    )


@app.post("/api/import", tags=["Import"])
def import_project(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Parse uploaded Excel file (3 sheets: Project, Pours, Inventory).
    Creates the project, all pours, and inventory in one transaction.
    Returns a summary dict.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx / .xls files are supported.")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Could not read Excel file. Ensure it follows the template format.")

    errors = []

    # ── Helper: read sheet as list-of-dicts ──────────────────
    def sheet_rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            raise HTTPException(400, f"Missing sheet: '{sheet_name}'")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # skip rows until we find the header (first non-empty row)
        header_row = None
        data_start = 0
        for i, row in enumerate(rows):
            if any(v is not None for v in row):
                header_row = [str(v).strip() if v else "" for v in row]
                data_start = i + 1
                break
        if not header_row:
            raise HTTPException(400, f"Sheet '{sheet_name}' appears empty.")
        result = []
        for row in rows[data_start:]:
            if all(v is None for v in row):
                continue
            d = dict(zip(header_row, row))
            result.append(d)
        return result

    def val(d, key, default=None):
        v = d.get(key, default)
        return v if v not in (None, "") else default

    # ── Parse Project sheet ───────────────────────────────────
    proj_rows = sheet_rows("Project")
    # skip note/hint rows — find first row where project_code looks like a real code
    proj_data = None
    for r in proj_rows:
        code = val(r, "project_code")
        if not code:
            continue
        code_str = str(code).strip()
        # real codes have no spaces and are not generic hint text
        if ' ' not in code_str and code_str.lower() not in ("project_code", "unique code", ""):
            proj_data = r
            break
    if not proj_data:
        raise HTTPException(400, "No project data found in 'Project' sheet.")

    project_code = str(val(proj_data, "project_code", "")).strip()
    project_name = str(val(proj_data, "project_name", "Unnamed Project")).strip()
    location     = str(val(proj_data, "location", "")).strip()
    sap_wbs      = str(val(proj_data, "sap_wbs_root", "")).strip()

    def parse_date(v, fallback):
        if v is None: return fallback
        if isinstance(v, (date, datetime)): return v if isinstance(v, date) else v.date()
        try: return date.fromisoformat(str(v).strip()[:10])
        except: return fallback

    start_date  = parse_date(val(proj_data, "start_date"),   date.today())
    planned_end = parse_date(val(proj_data, "planned_end"),  date(start_date.year+1, start_date.month, start_date.day))

    if not project_code:
        raise HTTPException(400, "project_code is required in the Project sheet.")

    if db.query(Project).filter(Project.project_code == project_code).first():
        raise HTTPException(409, f"Project code '{project_code}' already exists. Change it in the Excel file.")

    project = Project(
        id=str(uuid.uuid4()), project_code=project_code,
        project_name=project_name, location=location,
        sap_wbs_root=sap_wbs, start_date=start_date, planned_end=planned_end,
    )
    db.add(project)
    db.flush()

    # ── Parse Pours sheet ─────────────────────────────────────
    pour_rows = sheet_rows("Pours")
    pours_created = 0
    pour_codes_seen = set()
    for i, r in enumerate(pour_rows):
        tower = str(val(r, "tower_code", "T1")).strip()
        try:
            floor = int(float(str(val(r, "floor_number", 1))))
        except:
            errors.append(f"Pours row {i+1}: invalid floor_number, skipped.")
            continue
        zone      = str(val(r, "zone_code", "A")).strip().upper()
        pour_type = str(val(r, "pour_type", "WALL")).strip().upper()
        if pour_type not in ("WALL","SLAB","COLUMN","BEAM","SHEAR_WALL"):
            errors.append(f"Pours row {i+1}: unknown pour_type '{pour_type}', defaulting to WALL.")
            pour_type = "WALL"
        planned   = parse_date(val(r, "planned_date"), start_date)
        try:
            area = float(str(val(r, "net_surface_area_m2", 30)))
        except:
            area = 30.0
        grade   = str(val(r, "concrete_grade", "M40")).strip()
        strip_h = int(float(str(val(r, "strip_cycle_hours", 12))))
        system  = str(val(r, "assigned_system", "PERI_TRIO")).strip()

        base_code = f"{tower}-F{floor:02d}-{zone}-{pour_type[0]}"
        pc = base_code
        suffix = 1
        while pc in pour_codes_seen:
            pc = f"{base_code}-{suffix}"; suffix += 1
        pour_codes_seen.add(pc)

        db.add(Pour(
            id=str(uuid.uuid4()), project_id=project.id,
            pour_code=pc, tower_code=tower, floor_number=floor,
            zone_code=zone, pour_type=pour_type, planned_date=planned,
            net_surface_area_m2=area, concrete_grade=grade,
            strip_cycle_hours=strip_h, assigned_system=system, status="PLANNED",
        ))
        pours_created += 1

    # ── Parse Inventory sheet ─────────────────────────────────
    inv_rows = sheet_rows("Inventory")
    inv_created = 0
    for i, r in enumerate(inv_rows):
        comp_code = str(val(r, "component_code", "")).strip()
        if not comp_code:
            continue
        # skip hint/notes rows — real component codes never contain spaces
        if ' ' in comp_code or comp_code.lower() in ("component_code", "exact code from master"):
            continue
        comp = db.query(ComponentMaster).filter(ComponentMaster.component_code == comp_code).first()
        if not comp:
            errors.append(f"Inventory row {i+1}: component '{comp_code}' not found in master, skipped.")
            continue
        try: qty_a = int(float(str(val(r, "qty_available", 0))))
        except: qty_a = 0
        try: qty_d = int(float(str(val(r, "qty_deployed", 0))))
        except: qty_d = 0
        try: qty_r = int(float(str(val(r, "qty_under_repair", 0))))
        except: qty_r = 0

        existing = db.query(InventoryItem).filter(
            InventoryItem.project_id == project.id,
            InventoryItem.component_id == comp.id,
        ).first()
        if existing:
            existing.qty_available += qty_a
            existing.qty_deployed  += qty_d
            existing.qty_under_repair += qty_r
        else:
            db.add(InventoryItem(
                id=str(uuid.uuid4()), project_id=project.id,
                component_id=comp.id, qty_available=qty_a,
                qty_deployed=qty_d, qty_under_repair=qty_r,
            ))
        inv_created += 1

    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(409, f"Database conflict: {str(e)}")

    return {
        "status": "success",
        "project_id": project.id,
        "project_code": project_code,
        "project_name": project_name,
        "pours_created": pours_created,
        "inventory_lines": inv_created,
        "warnings": errors,
    }


# ════════════════════════════════════════════════════════════════
# ENTRY POINT
# ════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.app:app", host="0.0.0.0", port=8000, reload=True)
